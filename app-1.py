from flask import Flask, render_template, request, redirect, url_for
import openpyxl
from openpyxl import Workbook
import os
from datetime import datetime

app = Flask(__name__)

# Initialize the Excel file if it doesn't exist
def init_excel():
    file_path = r'C:\Users\Deepa\Downloads\rtr\faculty_events.xlsx'
    
    if os.path.exists(file_path):
        try:
            wb = openpyxl.load_workbook(file_path)
            return
        except Exception as e:
            print(f"Error loading workbook: {e}. Recreating the file.")
            os.remove(file_path)  # Remove corrupted file
    
    # Create a new workbook with department sheets and headers
    wb = Workbook()
    departments = ['MCT', 'AIDS', 'AIML', 'AERO', 'BIOTECH', 'AUTO', 'BME', 'CHEMICAL', 
                   'CIVIL', 'CSE', 'CSBS', 'CSD', 'CYBER SECURITY', 'ECE', 'EEE', 'FT', 
                   'CHEMISTRY', 'ENGLISH', 'MATHEMATICS', 'PHYSICS', 'IT', 'MANAGEMENT', 
                   'MECH', 'R&A']
    
    headers = [
        'Faculty ID', 
        'Name',
        'Designation',  
        'Event Type (FDP/Workshop/Seminar/Course)',
        'Participation Type (Attended/Organized)', 
        'Title of the Event / Course',
        'Month & Year', 
        'Event Date (From)',  
        'Event Date (To)',    
        'Total Days',
        'Mode (Online/Offline)',
        'Number of Participants', 
        'Organization Name', 
        'Sponsoring Agency',
        'Certificate Upload (YES/NO)', 
        'Report on Learning Outcomes Upload (YES/NO)',
        'Proceedings/Study Materials Upload (YES/NO)',
        'Report on Learning Outcomes File',  # New field for file path
    ]
    
    for department in departments:
        sheet = wb.create_sheet(department)
        sheet.append(headers)
    
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    
    wb.save(file_path)

# Route for faculty event form
@app.route('/')
def faculty_event_form():
    return render_template('faculty_event_form-1.html')

# Route for form submission
@app.route('/submit', methods=['POST'])
def submit():
    if request.method == 'POST':
        # Fetch form data
        data = {
            'faculty_id': request.form['faculty_id'],
            'name': request.form['name'],
            'designation': request.form['designation'],
            'department': request.form['department'],
            'event_type': request.form['event_type'],
            'participation_type': request.form['participation_type'],
            'event_title': request.form['event_title'],
            'month_year': request.form['month_year'],
            'event_dates_from': request.form['event_dates_from'],  
            'event_dates_to': request.form['event_dates_to'],     
            'mode': request.form['mode'],
            'organization': request.form['organization'],
            'sponsoring_agency': request.form['sponsoring_agency'],
            'certificate_upload': request.form['certificate_upload'],
            'Report_on_Learning_Outcomes_Upload': request.form['Report_on_Learning_Outcomes_Upload'],
            'proceedings_upload': request.form['proceedings_upload']
        }

        # Handle optional number of participants
        if data['participation_type'] == 'Organized':
            data['num_participants'] = request.form['num_participants']
        else:
            data['num_participants'] = ''  # Empty if not provided

        # Calculate total days between event dates
        date_format = "%Y-%m-%d"  
        start_date = datetime.strptime(data['event_dates_from'], date_format)
        end_date = datetime.strptime(data['event_dates_to'], date_format)
        total_days = (end_date - start_date).days  # Calculate difference in days

        # Load the existing Excel file
        wb = openpyxl.load_workbook('faculty_events.xlsx')
        sheet = wb[data['department']]  # Select the sheet based on department

        # Prepare row data
        row_data = [
            data['faculty_id'], 
            data['name'],
            data['designation'], 
            data['event_type'], 
            data['participation_type'],
            data['event_title'], 
            data['month_year'], 
            data['event_dates_from'],  
            data['event_dates_to'],    
            total_days,
            data['mode'],
            data['num_participants'],  
            data['organization'], 
            data['sponsoring_agency'],
            data['certificate_upload'], 
            data['Report_on_Learning_Outcomes_Upload'], 
            data['proceedings_upload'],
            ""  # No file path for report (file upload is removed)
        ]

        # Append row to the appropriate department sheet
        sheet.append(row_data)
        
        # Save the workbook
        wb.save('faculty_events.xlsx')

        # Redirect with a success message
        return redirect(url_for('faculty_event_form', message='Event Added Successfully!'))

if __name__ == '__main__':
    init_excel()
    app.run(debug=True)
